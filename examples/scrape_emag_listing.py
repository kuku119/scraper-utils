"""
爬取 Emag 搜索页的 listing
"""

# TODO 不想做了

import asyncio
from collections import defaultdict
import itertools
from random import randint, uniform
from pathlib import Path
from time import perf_counter
from typing import Any
from zipfile import ZipFile

from loguru import logger
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment
from playwright.async_api import Page

from scraper_utils.constants.time_constant import MS1000
from scraper_utils.enums.browser_enum import ResourceType as RT
from scraper_utils.utils.browser_util import launch_persistent_browser, create_new_page, close_browser
from scraper_utils.utils.emag_url_util import build_search_url, build_search_urls, clean_product_image_url
from scraper_utils.utils.json_util import write_json_async, read_json_async, json_loads
from scraper_utils.utils.time_util import now_str
from scraper_utils.utils.workbook_util import read_workbook_async, write_workbook_async


def backup_listings_json_file(CWD: Path, json_save_dir: Path):
    """把上一次爬取的 json 保存成 zip 文件"""
    backup_json_files = list(_ for _ in json_save_dir.glob('*.json'))
    if len(backup_json_files) > 0:
        with ZipFile(CWD.joinpath(f'temp/emag_jsons/backup.{now_str('%Y%m%d_%H%M%S')}.zip'), 'w') as zfp:
            for file in json_save_dir.glob('*.json'):
                zfp.write(file, arcname=file.name)
                file.unlink(missing_ok=True)


async def search_favorite_listings(CWD: Path, json_save_dir: Path, target_rows: list):
    """用标签搜索产品"""
    await launch_persistent_browser(
        executable_path=r'C:\Program Files\Google\Chrome\Application\chrome.exe',
        user_data_dir=CWD.joinpath('temp/chrome_data'),
        channel='chrome',
        headless=False,
        timeout=60 * MS1000,
    )
    abort_resources = (RT.MEDIA, RT.IMAGE)  # 屏蔽图片、音视频资源

    workbook = await read_workbook_async(CWD.joinpath('temp/emag_keyword.xlsx'), read_only=True)
    worksheet = workbook.active

    for row in target_rows:  # 遍历标签表格
        keyword = str(worksheet.cell(row, 4).value)
        search_url = build_search_url(keyword=keyword, page=1)
        logger.info(f'爬取关键词：[row] {keyword}')

        page = await create_new_page(stealth_page=True, abort_resources=abort_resources)
        await page.goto(search_url, timeout=60 * MS1000)

        listings = await parse_favorite_listings(page=page)
        await write_json_async(
            file=json_save_dir.joinpath(f'{row}.json'),
            data={
                'keyword': keyword,
                'listings': listings,
            },
            indent=4,
        )

        await asyncio.sleep(randint(20, 40))  # 随机延时
        await page.close()

    await close_browser()


async def parse_favorite_listings(page: Page) -> list[str]:
    """解析搜索页上的带 Top favorite 的 listing"""
    logger.debug(f'解析：{page.url}')

    result: list[str] = list()

    wheel_start_time = perf_counter()
    while True:  # 模拟鼠标滚轮向下滚动网页，直至 item_card_tag 数量达标或者时间超时
        item_card_tags_1 = page.locator(r'//div[@class="card-item card-standard js-product-data js-card-clickable "]')
        item_card_tags_2 = page.locator(r'//div[@class="card-item card-fashion js-product-data js-card-clickable"]')

        item_card_tag_count = await item_card_tags_1.count() + await item_card_tags_2.count()
        if item_card_tag_count >= 64:
            break

        await page.mouse.wheel(delta_y=randint(50, 150), delta_x=0)
        await page.wait_for_timeout(uniform(0, 0.5) * MS1000)
        if perf_counter() - wheel_start_time >= 30:  # 30 秒后数量还不够，就有多少爬多少（有些关键词的搜索结果就那么多）
            break

    logger.debug(f'定位到 {item_card_tag_count} 个 item_card_tag')
    for item_card_tag in await item_card_tags_1.all() + await item_card_tags_2.all():
        top_favourite_tag = item_card_tag.locator(r'//span[text()="Top Favorite"]')
        if await top_favourite_tag.count() > 0:
            item_title_tag = item_card_tag.first.locator(r'//a[@data-zone="title"]')
            if await item_title_tag.count() > 0:
                item_title = await item_title_tag.first.inner_text(timeout=MS1000)
                result.append(item_title)

    logger.debug(f'找到 {len(result)} 个符合条件的 listing')
    return result


async def concat_favorite_listings(CWD: Path, json_save_dir: Path):
    """合并 Top favorite 的爬取结果"""
    red_bold_font = Font(color='FF0000', bold=True, size=14)  # 红字、加粗、14 号字
    yellow_bg = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # 黄底
    align = Alignment(wrap_text=True, horizontal='center', vertical='center')  # 居中、自动换行

    workbook = Workbook()
    sheet = workbook.active

    # 写入数据
    files = sorted(list(json_save_dir.glob('*.json')), key=lambda p: int(p.stem))
    for i, f in enumerate(files):
        data: dict[str, str | list[str]] = await read_json_async(file=f)
        keyword: str = data['keyword']
        listings: list[str] = data['listings']

        sheet.cell(row=1, column=i + 1, value=keyword)
        sheet.cell(row=1, column=i + 1).font = red_bold_font
        sheet.cell(row=1, column=i + 1).fill = yellow_bg

        for row, listing in enumerate(listings, start=2):
            sheet.cell(row=row, column=i + 1, value=listing)

    # 设置各列宽度为 50
    for col in sheet.columns:
        col_letter = col[0].column_letter
        sheet.column_dimensions[col_letter].width = 30

    # 设置各单元格居中
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = align

    result_path = await write_workbook_async(file=CWD.joinpath('temp/emag_listings.xlsx'), workbook=workbook)
    logger.success(f'程序结束，结果保存至：{result_path}')


def backup_keyword_json_file(CWD: Path, json_save_dir: Path):
    """"""
    # TODO


async def search_keywords(CWD: Path, json_save_dir: Path):
    """用关键词去搜索产品，爬取所有前 3 页的所有产品，不管有无 top"""

    workbook = await read_workbook_async(file=CWD.joinpath('temp/emag关键词.xlsx'), read_only=True)
    worksheet = workbook.active

    abort_resources = (RT.MEDIA, RT.IMAGE)
    browser = await launch_persistent_browser(
        executable_path=r'C:\Program Files\Google\Chrome\Application\chrome.exe',
        user_data_dir=CWD.joinpath('temp/chrome_data'),
        channel='chrome',
        headless=True,
        timeout=60 * MS1000,
    )

    for row in range(1, worksheet.max_row + 1):
        # for row in range(3, 4):  # 测试
        keyword = worksheet.cell(row, 2).value
        if keyword is None or keyword.startswith('工') or keyword.startswith('关'):
            continue

        logger.debug(f'读取关键词 "{keyword}"')

        result: dict[str, list[str]] = {
            'keyword': keyword,
            'products': list(),
        }
        for p, url in enumerate(build_search_urls(keyword=keyword, max_page=2)):
            # for p, url in enumerate(build_search_urls(keyword=keyword, max_page=1), start=1):  # 测试
            page = await create_new_page(stealth_page=True, abort_resources=abort_resources)
            await page.goto(url, timeout=60 * MS1000)

            result_one_page = await parse_keyword_search_result(page)
            for product_data in result_one_page:
                result['products'].append(product_data)
            await write_json_async(file=json_save_dir.joinpath(f'{row}-{p}.json'), data=result, indent=4)

            await asyncio.sleep(randint(20, 40))  # 随机延时
            await page.close()

    await close_browser()


async def parse_keyword_search_result(page: Page) -> list[dict[str, Any]]:
    """解析关键词搜索页，解析产品 url 就行"""
    logger.debug(f'解析页面 "{page.url}"')

    result_total: list[dict[str, Any]] = list()

    wheel_start_time = perf_counter()
    while True:  # 模拟鼠标滚轮向下滚动网页，直至 item_card_tag 数量达标或者时间超时
        item_card_tags_1 = page.locator(r'//div[@class="card-item card-standard js-product-data js-card-clickable "]')
        item_card_tags_2 = page.locator(r'//div[@class="card-item card-fashion js-product-data js-card-clickable"]')

        item_card_tag_count = await item_card_tags_1.count() + await item_card_tags_2.count()
        if item_card_tag_count >= 64:
            break

        await page.mouse.wheel(delta_y=randint(100, 500), delta_x=0)
        await page.wait_for_timeout(uniform(0, 0.5) * MS1000)
        if perf_counter() - wheel_start_time >= 30:  # 30 秒后数量还不够，就有多少爬多少（有些关键词的搜索结果就那么多）
            break

    logger.debug(f'定位到 {item_card_tag_count} 个 item_card_tag')
    for item_card_tag in await item_card_tags_1.all() + await item_card_tags_2.all():
        result_one_item: dict[str, Any] = dict()

        data_product_button_tag = item_card_tag.locator(r'//button[@data-product!=""]')
        if await data_product_button_tag.count() > 0:
            product_data = await data_product_button_tag.first.get_attribute('data-product')
            if product_data is not None:
                result_one_item.update(json_loads(product_data))

        # 是否要忽略 bundle-image? //div[@class="bundle-image position-static"] 目前是忽略的
        product_image_tag = item_card_tag.locator(
            r'//div[@class="img-component position-relative card-v2-thumb-inner"]/img[@src!=""]'
        )
        if await product_image_tag.count() > 0:
            product_image_url = await product_image_tag.first.get_attribute('src', timeout=MS1000)
            if product_image_url is not None:
                product_image_url = clean_product_image_url(url=product_image_url)
                result_one_item['image_url'] = product_image_url

        result_total.append(result_one_item)

    return result_total


async def concat_keyword_search_result(CWD: Path, json_save_dir: Path):
    """合并关键词搜索结果"""
    # TODO
    red_bold_font = Font(color='FF0000', bold=True, size=14)  # 红字、加粗、14 号字
    yellow_bg = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # 黄底

    workbook = Workbook()
    workbook.remove(workbook.active)

    files = list(json_save_dir.glob('*.json'))
    files_group: dict[int, list[Path]] = defaultdict(list)
    for f in files:  # 分组
        files_group[int(f.stem.split('-')[0])].append(f)
    for g in files_group:  # 组内排序
        files_group[g].sort(key=lambda p: int(p.stem.split('-')[1]))

    for i, k in enumerate(sorted(files_group), start=1):
        products_one_keyword: list[dict[str, Any]] = list()
        for f in files_group[k]:
            data: dict[str, str | list[dict[str, Any]]] = await read_json_async(file=f)
            keyword: str = data['keyword']
            products_one_json: list[dict[str, Any]] = data['products']
            products_one_keyword.append(products_one_json)
        sheet: Worksheet = workbook.create_sheet(str(k), index=i)

        sheet.cell(1, 1, value=keyword)
        sheet.cell(1, 1).font = red_bold_font
        sheet.cell(1, 1).fill = yellow_bg

        for idx, product in enumerate(products_one_keyword):
            puk = product.get('puk', None)
            product_name = product.get('product_name')
            sheet.cell(
                idx + 2,
            )


if __name__ == '__main__':
    CWD = Path.cwd()
    listings_json_dir = CWD.joinpath('temp/emag_jsons')
    keyword_json_dir = CWD.joinpath('temp/emag_keyword_jsons')

    async def main():
        """"""
        backup_listings_json_file(CWD, json_save_dir=listings_json_dir)
        try:
            await search_favorite_listings(CWD, json_save_dir=listings_json_dir, target_rows=list(range(4, 62)))
        except Exception as e:
            logger.error(e)
        except BaseException as be:
            logger.error(be)
        await concat_favorite_listings(CWD, json_save_dir=listings_json_dir)

        backup_keyword_json_file(CWD, json_save_dir=keyword_json_dir)
        try:
            await search_keywords(CWD, json_save_dir=keyword_json_dir)
        except Exception as e:
            logger.error(e)
        except BaseException as be:
            logger.error(be)
        await concat_keyword_search_result(CWD, json_save_dir=keyword_json_dir)

    asyncio.run(main())
