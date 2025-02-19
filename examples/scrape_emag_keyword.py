"""
根据关键词爬取 emag 搜索页的产品 url
"""

import asyncio
import re
import sys
from asyncio import Semaphore
from asyncio.exceptions import CancelledError
from dataclasses import dataclass, field
from pathlib import Path
from random import randint, uniform
from re import Pattern
from time import perf_counter
from typing import Generator, Iterable, Optional, Self

from aiohttp import ClientResponseError, ClientSession, ClientTimeout
from fake_http_header import FakeHttpHeader
from loguru import logger
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.fills import PatternFill
from openpyxl.styles.fonts import Font
from openpyxl.workbook import Workbook
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.worksheet import Worksheet
from playwright.async_api import Page

from scraper_utils.constants.time_constant import MS1000
from scraper_utils.enums.browser_enum import ResourceType
from scraper_utils.exceptions.browser_exception import PlaywrightError
from scraper_utils.utils.browser_util import (
    launch_persistent_browser,
    close_browser,
    create_new_page,
)
from scraper_utils.utils.emag_url_util import (
    build_search_urls,
    build_product_url,
    clean_product_image_url,
)
from scraper_utils.utils.file_util import write_bytes_async, path_exists
from scraper_utils.utils.image_util import read_image_sync, resize_image
from scraper_utils.utils.workbook_util import (
    read_workbook_sync,
    write_workbook_sync,
    string_column_to_integer_column as s2i,
    insert_image,
)

logger.remove()
logger.add(
    sys.stderr,
    format=(
        '[<green>{time:HH:mm:ss}</green>] [<level>{level}</level>]\t'
        '[<cyan>{name}</cyan>:<cyan>{function}</cyan>:<cyan>{line}</cyan>] >>> '
        '<level>{message}</level>'
    ),
)


cur_work_dir = Path().cwd()

product_image_ext_pattern = re.compile(r'/images/[0-9a-z_]+\.([a-z]+)')  # 获取图片拓展名


@dataclass
class ItemCardParseResult:
    """搜索结果页上，单个产品标签所解析出的产品信息"""

    pnk: str
    top_favorite: bool = False  # 是否带有 top 标志
    image_url: Optional[str] = None  # 产品图 url
    review_count: Optional[int] = None  # 评论数
    price: Optional[float] = None  # 价格
    #
    __url: Optional[str] = None
    __origin_image_url: Optional[str] = None
    __image_ext: Optional[str] = None
    __image_save_path: Optional[Path] = None

    def __hash__(self):
        return hash(self.pnk)

    def __eq__(self, other: Self):
        if isinstance(other, self.__class__):
            return self.pnk == other.pnk
        return False

    @property
    def url(self) -> str:
        """产品详情页 url"""
        if self.__url is not None:
            return self.__url

        self.__url = build_product_url(pnk=self.pnk)
        return self.__url

    @property
    def origin_image_url(self) -> Optional[str]:
        """产品图原图 url"""
        if self.__origin_image_url is not None:
            return self.__origin_image_url

        if self.image_url is not None:
            self.__origin_image_url = clean_product_image_url(url=self.image_url)
            return self.__origin_image_url

        return None

    @property
    def image_ext(self) -> Optional[str]:
        """产品图的文件拓展名"""
        if self.__image_ext is not None:
            return self.__image_ext

        if self.origin_image_url is not None:
            image_ext_match = product_image_ext_pattern.search(self.origin_image_url)
            if image_ext_match is not None:
                image_ext = image_ext_match.group(1)
                if image_ext is None or len(image_ext) == 0:
                    raise ValueError('image_ext 为空')
                if image_ext.lower() == 'jpg':
                    image_ext = 'jpeg'
                self.__image_ext = image_ext
                return self.__image_ext

        return None

    @property
    def image_save_path(self) -> Optional[Path]:
        """产品图下载后保存的位置"""
        if self.__image_save_path is not None:
            return self.__image_save_path

        if self.image_ext is not None:
            self.__image_save_path = cur_work_dir.joinpath(
                f'temp/emag_product_images/{self.pnk}.{self.image_ext}'
            )
            return self.__image_save_path

        return None


@dataclass
class OneKeywordSearchResult:
    """单个关键词所包含的爬取结果"""

    keyword: str
    products: list[ItemCardParseResult] = field(default_factory=list)

    def __add__(self, other: Self):
        """合并两个结果，并去除其中的重复产品"""
        if not isinstance(other, self.__class__):
            raise NotImplementedError(f'无法将 {type(self)} 和 {type(other)} 进行合并')

        if self.keyword != other.keyword:
            raise ValueError('两者的关键词不同，无法合并')
        return self.__class__(
            keyword=self.keyword,
            products=list(dict.fromkeys(self.products + other.products)),
        )


async def start_scrape(cwd: Path):
    """打开浏览器，初始化页面，读取关键词，根据关键词生成搜索页，生成加载完的页面"""
    # 存放关键词的工作簿
    keyword_workbook_path = cwd.joinpath('temp/scrape_emag_keyword_示例关键词.xlsx')
    # 要爬取的关键词的行号
    start_row = 123
    end_row = 149
    # 测试模式
    test_mode = False

    # 存储爬取结果的工作簿
    result_workbook = Workbook()
    # 移除第一个工作表
    result_workbook.remove(result_workbook.active)

    # 解析页面用的正则表达式
    pnk_pattern = re.compile(r'/pd/([0-9A-Z]{9})(/|$)')  # 解析 pnk
    review_count_pattern = re.compile(r'\((\d+)\)')  # 解析 评论数
    price_pattern = re.compile(r'[^\d,]')  # 解析 价格

    # 保存所有关键词的搜索结果
    total_result: list[OneKeywordSearchResult] = list()

    # 遍历关键词进行爬取
    row = -1
    try:
        for row, keyword in load_keywords(keyword_workbook_path, range(start_row, end_row + 1)):
            one_keyword_result = OneKeywordSearchResult(keyword=keyword)  # 单个关键词的爬取结果
            logger.info(f'开始爬取第 {row} 行关键词 "{keyword}"')
            # 爬取关键词搜索结果
            for url in build_search_urls(keyword=keyword, max_page=1):
                page = await create_new_page()
                try:
                    await page.goto(url, timeout=60 * MS1000)
                except PlaywrightError as pe:
                    logger.error(pe)
                else:
                    one_page_result = OneKeywordSearchResult(
                        keyword=keyword,
                        products=await parse_search_page(
                            page=page,
                            pnk_pattern=pnk_pattern,
                            review_count_pattern=review_count_pattern,
                            price_pattern=price_pattern,
                        ),
                    )  # 单个页面的爬取结果
                    one_keyword_result = one_keyword_result + one_page_result
                finally:
                    if test_mode is False:
                        # 每爬完一个 url 就随机等待 30-60 秒
                        await asyncio.sleep(30 + randint(0, 30))
                    await page.close()
            total_result.append(one_keyword_result)

            # 并发下载产品图
            await download_keyword_result_images(one_keyword_result)

            # 将单个关键词的爬取结果保存成工作表
            one_keyword_result_sheet: Worksheet = result_workbook.create_sheet(
                keyword if len(keyword) < 28 else keyword[:28] + '...'
            )
            save_keyword_result(one_keyword_result_sheet, one_keyword_result)

            logger.success(f'"{keyword}" 爬取结束')

    # Ctrl + C 不会退出程序，而是停止爬取并运行下面的结果保存
    except CancelledError:
        logger.warning(f'CancelledError: 爬取任务被中断，当前爬取到 {row}')
    except KeyboardInterrupt:
        logger.warning(f'KeyboardInterrupt: 爬取任务被中断，当前爬取到 {row}')
    finally:
        result_path = write_workbook_sync(
            file=cwd.joinpath(
                f'temp/scrape_emag_keyword_结果_{start_row}-{min(row, end_row)}.xlsx'
            ),
            workbook=result_workbook,
        )
        logger.success(f'程序结束，结果已保存至 "{result_path}"')


def load_keywords(
    workbook_path: Path,
    target_rows: Iterable[int],
) -> Generator[tuple[int, str]]:
    """从关键词工作簿读取关键词"""
    logger.info(f'读取关键词表格 "{workbook_path}"')

    workbook = read_workbook_sync(file=workbook_path, read_only=True)
    sheet: Worksheet = workbook.active

    for row in target_rows:
        keyword: Optional[str] = sheet.cell(row, 1).value
        if keyword is None:
            continue
        yield row, str(keyword)


async def parse_search_page(
    page: Page,
    pnk_pattern: Pattern[str],
    review_count_pattern: Pattern[str],
    price_pattern: Pattern[str],
) -> list[ItemCardParseResult]:
    """解析页面"""
    logger.debug(f'解析页面 "{page.url}"')

    # 模拟鼠标滚轮向下滚动网页，直至 item_card_tag 数量达标或者时间超时
    wheel_start_time = perf_counter()
    while True:
        item_card_tags_1 = page.locator(
            '//div[@class="card-item card-standard js-product-data js-card-clickable " and @data-url!=""]'
        )
        item_card_tags_2 = page.locator(
            '//div[@class="card-item card-fashion js-product-data js-card-clickable" and @data-url!=""]'
        )

        item_card_tag_count = await item_card_tags_1.count() + await item_card_tags_2.count()
        if item_card_tag_count >= 72:  # 目标数量 72
            break

        # 随机向下滚动
        await page.mouse.wheel(delta_y=randint(500, 1000), delta_x=0)
        await page.wait_for_timeout(uniform(0, 0.5) * MS1000)

        # 一定时间后后数量还不够，就有多少爬多少（有些关键词的搜索结果就那么多）
        if perf_counter() - wheel_start_time >= 10:  # 10 秒
            break
    logger.debug(f'定位到 {item_card_tag_count} 个 card-item 标签')

    ##########

    # 提取各个 item_card_tag 的产品 pnk 并判断其是否有 top 标
    result: list[ItemCardParseResult] = list()
    for item_card_tag in await item_card_tags_1.all() + await item_card_tags_2.all():
        # 查找 pnk
        data_url: str = await item_card_tag.get_attribute('data-url', timeout=MS1000)
        pnk_match = pnk_pattern.search(data_url)
        if pnk_match is not None:
            pnk = pnk_match.group(1)  # 1. 产品 pnk

            # 判断是否有 top 标
            top_favorite_tag = item_card_tag.locator('//span[text()="Top Favorite"]')
            top_favorite = await top_favorite_tag.count() > 0  # 2. 是否有 Top Favorite 标志

            one_item = ItemCardParseResult(pnk=pnk, top_favorite=top_favorite)

            # 获取产品图 url
            image_url: Optional[str] = None  # 3. 产品图 url
            image_tag = item_card_tag.locator(
                '//div[@class="img-component position-relative card-v2-thumb-inner"]/img[@src!=""]'
            )
            if await image_tag.count() > 0:
                image_url = await image_tag.first.get_attribute('src', timeout=MS1000)
            if image_url is not None:
                image_url = clean_product_image_url(url=image_url)
            one_item.image_url = image_url

            # 获取评论数
            review_count: Optional[int] = None  # 4. 评论数
            review_count_tag = item_card_tag.locator(
                '//div[@class="star-rating-text "]/span[@class="visible-xs-inline-block " and text()!=""]'
            )
            if await review_count_tag.count() > 0:
                review_count_text = await review_count_tag.first.inner_text(timeout=MS1000)
                review_count_match = review_count_pattern.search(review_count_text)
                if review_count_match is not None:
                    review_count = int(review_count_match.group(1))
            one_item.review_count = review_count

            # 获取价格
            price: Optional[float] = None  # 5. 价格
            price_tag = item_card_tag.locator('//p[@class="product-new-price"]')
            if await price_tag.count() > 0:
                price_text = await price_tag.first.inner_text(timeout=MS1000)
                if len(price_text) > 0:
                    price = float(price_pattern.sub('', price_text).replace(',', '.'))
            one_item.price = price

            result.append(one_item)

    logger.debug(f'解析出 {len(result)} 个产品')

    return result


async def download_keyword_result_images(
    keyword_result: OneKeywordSearchResult,
    concurrent_limit: int = 8,
):
    """下载单个关键词的产品图到本地"""
    logger.info(f'下载 "{keyword_result.keyword}" 的产品图')

    fake_header = FakeHttpHeader(browser='chrome').as_header_dict()
    concurrent_semaphore = Semaphore(concurrent_limit)

    # 下载产品图的任务
    async with ClientSession() as client_session:
        download_image_tasks = [
            download_image_task(
                image_url=pr.origin_image_url,
                save_path=pr.image_save_path,
                client_session=client_session,
                fake_header=fake_header,
                semaphore=concurrent_semaphore,
            )
            for pr in keyword_result.products
            if pr.image_save_path is not None
        ]
        await asyncio.gather(*download_image_tasks)

    logger.success(f'"{keyword_result.keyword}" 的产品图下载结束')


async def download_image_task(
    image_url: str,
    save_path: Path,
    client_session: ClientSession,
    fake_header: dict[str, str],
    semaphore: Semaphore,
) -> None:
    """下载单个产品图"""
    if path_exists(path=save_path):
        # logger.debug(f'产品图已存在 "{save_path}"')
        pass
    else:
        async with semaphore:
            try:
                logger.debug(f'下载产品图 "{image_url}"')
                async with client_session.get(
                    url=image_url,
                    headers=fake_header,
                    timeout=ClientTimeout(total=30),
                ) as response:
                    response.raise_for_status()
                    content = await response.read()
                    await write_bytes_async(file=save_path, data=content)

            except ClientResponseError as cre:
                logger.error(f'下载产品图时出错\n{cre}')
            except CancelledError as ce:
                logger.error(f'下载产品图时出错\n{ce}')
            except TimeoutError as te:
                logger.error(f'下载产品图时出错\n{te}')


def save_keyword_result(
    sheet: Worksheet,
    keyword_result: OneKeywordSearchResult,
) -> None:
    """保存单个关键词的爬取结果到工作表中"""
    logger.info(f'保存 "{keyword_result.keyword}" 的爬取结果为工作表')

    # 特定单元格的样式
    red_bold_font = Font(color='FF0000', bold=True)  # 红字、加粗，标题栏用
    yellow_fill_color = PatternFill(fill_type='solid', fgColor='FFFF00')  # 黄色填充色，标题栏用
    hyperlink_font = Font(color='0000FF', underline='single')  # 蓝字、下划线，超链接单元格用
    center_align = Alignment(
        wrap_text=True, horizontal='center', vertical='center'
    )  # 居中、自动换行

    # 每爬完一个关键词就写入一个工作表
    keyword = keyword_result.keyword
    sheet['A1'] = '关键词'
    sheet['B1'] = keyword

    # 创建标题行
    sheet['A2'] = '产品链接'
    sheet['B2'] = '产品图'
    sheet['C2'] = '是否有 Top Favorite'
    sheet['D2'] = '评论数'
    sheet['E2'] = '价格'

    # 设置标题行的样式和各列的宽度
    for col in ['A', 'B', 'C', 'D', 'E']:
        sheet.cell(2, s2i(col)).alignment = center_align
        sheet.cell(2, s2i(col)).font = red_bold_font
        sheet.cell(2, s2i(col)).fill = yellow_fill_color
        sheet.column_dimensions[col].width = int(120 / 7)  # 16.04 字符 ≈ 100 磅

    # 遍历爬取结果，插入到工作表中
    for row, pr in enumerate(keyword_result.products, start=3):
        # 产品详情页 url
        sheet.cell(row=row, column=1, value=pr.url)
        sheet.cell(row=row, column=1).alignment = center_align
        sheet.cell(row=row, column=1).font = hyperlink_font
        sheet.cell(row=row, column=1).hyperlink = Hyperlink(ref=pr.url, target=pr.url)

        # 产品图
        if pr.origin_image_url is not None:
            sheet.cell(row=row, column=2, value=pr.origin_image_url)
            sheet.cell(row=row, column=2).alignment = center_align
            sheet.cell(row=row, column=2).hyperlink = Hyperlink(
                ref=pr.origin_image_url,
                target=pr.origin_image_url,
            )
            sheet.cell(row=row, column=2).font = hyperlink_font

            # 检查图片是否存在，存在则插入到工作表
            if path_exists(pr.image_save_path):
                sheet.row_dimensions[row].height = int(120 * 0.75)

                product_image = read_image_sync(file=pr.image_save_path)
                product_image = resize_image(image=product_image, height=120, width=120)
                insert_image(
                    sheet=sheet,
                    image=product_image,
                    row=row,
                    column='B',
                    image_format=pr.image_ext,
                )

        # Top Favorite
        sheet.cell(row=row, column=3, value=str(pr.top_favorite))
        sheet.cell(row=row, column=3).alignment = center_align

        # 评论数
        if pr.review_count is not None:
            sheet.cell(row=row, column=4, value=pr.review_count)
            sheet.cell(row=row, column=4).alignment = center_align

        # 价格
        if pr.price is not None:
            sheet.cell(row=row, column=5, value=pr.price)
            sheet.cell(row=row, column=5).alignment = center_align

    logger.success(f'"{keyword_result.keyword}" 的爬取结果保存结束')


if __name__ == '__main__':

    async def main():
        start_time = perf_counter()
        logger.info('程序启动')

        # 启动浏览器
        abort_res: tuple[ResourceType, ...] = (
            ResourceType.MEDIA,
            ResourceType.IMAGE,
            ResourceType.STYLESHEET,
            ResourceType.FONT,
        )
        await launch_persistent_browser(
            user_data_dir=cur_work_dir.joinpath('temp/chrome_data'),
            executable_path=r'C:\Program Files\Google\Chrome\Application\chrome.exe',
            channel='chrome',
            # headless=False,
            timeout=60 * MS1000,
            stealth=True,
            abort_resources=abort_res,
        )

        # 开始爬取
        await start_scrape(cwd=cur_work_dir)

        # 关闭浏览器
        await close_browser()

        end_time = perf_counter()
        logger.info(f'程序结束，用时 {round(end_time-start_time, 2)} 秒')

    asyncio.run(main())
