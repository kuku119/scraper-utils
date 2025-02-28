"""
爬取 eMAG 店铺
"""

import asyncio
import re
from random import uniform, randint
from pathlib import Path
from sys import stderr
from time import perf_counter
from typing import AsyncGenerator, Optional, Self, Generator

from aiohttp import ClientResponseError, ClientSession, ClientTimeout

from fake_http_header import FakeHttpHeader
from loguru import logger
from openpyxl.workbook import Workbook
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.worksheet import Worksheet
from playwright.async_api import BrowserContext, Page

from scraper_utils.constants.workbook_style import (
    HYPERLINK_FONT,
    RED_BOLD_FONT,
    YELLOW_FILL,
    TEXT_CENTER_WRAP_ALIGNMENT,
)
from scraper_utils.enums.browser_enum import ResourceType
from scraper_utils.utils.browser_util import PersistentContextManager, MS1000
from scraper_utils.utils.emag_util import build_product_url, clean_product_image_url
from scraper_utils.utils.file_util import write_file, select_file_dialog
from scraper_utils.utils.image_util import read_image, resize_image
from scraper_utils.utils.text_util import is_letter, is_number
from scraper_utils.utils.workbook_util import (
    insert_image,
    read_workbook,
    write_workbook,
    column_str2int as s2i,
)


# 当前工作目录
cwd = Path.cwd()


# 日志
logger.remove()
logger.add(
    stderr,
    format=(
        '[<green>{time:HH:mm:ss}</green>] [<level>{level:.3}</level>] '
        '[<cyan>{name}</cyan>:<cyan>{function}</cyan>:<cyan>{line}</cyan>] >>> '
        '<level>{message}</level>'
    ),
)


class CardItem:
    """单个 card-item 能解析出的产品信息"""

    def __init__(
        self,
        pnk: str,
        top_favorite: bool = False,
        thumb_image_url: Optional[str] = None,
        review_count: Optional[int] = None,
        price: Optional[float] = None,
    ):
        self.pnk = pnk
        self.top_favorite = top_favorite
        self.thumb_image_url = thumb_image_url
        self.review_count = review_count
        self.price = price

        # 缓存
        self.__url: Optional[str] = None
        self.__origin_image_url: Optional[str] = None
        self.__image_ext: Optional[str] = None
        self.__image_save_path: Optional[Path] = None

    def __hash__(self) -> int:
        return hash(self.pnk)

    def __eq__(self, other: object) -> bool:
        return isinstance(other, self.__class__) and self.pnk == other.pnk

    @property
    def url(self) -> str:
        """产品详情页链接"""
        if self.__url is not None:
            return self.__url
        self.__url = build_product_url(pnk=self.pnk)
        return self.__url

    @property
    def origin_image_url(self) -> Optional[str]:
        """产品图原图链接"""
        if self.__origin_image_url is not None:
            return self.__origin_image_url
        if self.thumb_image_url is not None:
            self.__origin_image_url = clean_product_image_url(url=self.thumb_image_url)
        return self.__origin_image_url

    @property
    def image_ext(self) -> Optional[str]:
        """产品图的文件拓展名"""
        if self.__image_ext is not None:
            return self.__image_ext
        if self.origin_image_url is not None:
            image_ext_match = re.search(r'/images/[0-9a-z_]+\.([a-z]+)', self.origin_image_url)
            if image_ext_match is not None:
                image_ext = str(image_ext_match.group(1))
                if len(image_ext) == 0:
                    raise ValueError(f'解析文件拓展名失败 "{self.origin_image_url}"')
                if image_ext.lower() == 'jpg':
                    image_ext = 'jpeg'
                self.__image_ext = image_ext
        return self.__image_ext

    @property
    def image_save_path(self) -> Optional[Path]:
        """产品图原图下载后保存的位置"""
        if self.__image_save_path is not None:
            return self.__image_save_path
        if self.image_ext is not None:
            self.__image_save_path = cwd.joinpath(f'emag_product_images/{self.pnk}.{self.image_ext}')
        return self.__image_save_path


class Shop:
    """单个店铺下的产品"""

    def __init__(self, url: str, shop_id: Optional[str] = None, products: Optional[list[CardItem]] = None) -> None:
        self.url = url
        self.__shop_id: Optional[str] = shop_id
        self.__products: list[CardItem] = products if products is not None else list()

    @property
    def products(self) -> list[CardItem]:
        """去除了重复产品的 products"""
        self.__products = list(dict.fromkeys(self.__products))
        return self.__products

    @property
    def shop_id(self) -> str:
        """从店铺链接中提取店铺编号"""
        if self.__shop_id is not None:
            return self.__shop_id

        shop_id_match = re.search(r'https://www.emag.ro/vendors/vendor/(.*?)($|/)', self.url)
        if shop_id_match is None or len(shop_id_match.group(1)) == 0:
            raise ValueError(f'无法解析出店铺编号 "{self.url}"')

        self.__shop_id = str(shop_id_match.group(1))
        return self.__shop_id

    def __add__(self, other) -> Self:
        """合并两个 Shop"""
        if not isinstance(other, self.__class__):
            return NotImplemented
        if self.shop_id != other.shop_id:
            raise ValueError(f'无法合并不同的店铺')
        return self.__class__(url=self.url, products=list(dict.fromkeys(self.products + other.products)))

    def append(self, product: CardItem) -> None:
        """添加产品"""
        self.__products.append(product)

    def __repr__(self) -> str:
        return f'{self.__class__.__name__}("{self.url}", {len(self.products)} products)'


async def start_scrape(browser_context: BrowserContext):
    """开始爬取"""

    # 存储爬取结果用的工作簿
    result_workbook = Workbook()
    # 去除第一个的空 Sheet
    result_workbook.remove(result_workbook.active)  # type: ignore

    # 匹配 card-item 标签
    card_item_selector = (
        '//div[(@class="card-item card-standard js-product-data js-card-clickable " or '
        '@class="card-item card-fashion js-product-data js-card-clickable") and @data-url!=""]'
    )

    try:
        for shop_url in load_shop_workbook():
            # 单个店铺的爬取结果
            one_shop_result = Shop(shop_url)

            new_page = await browser_context.new_page()
            await asyncio.sleep(10 + randint(0, 10))  # 随机等待一段时间
            await new_page.goto(shop_url, timeout=60 * MS1000)
            async for loaded_page in wait_page_load(new_page, card_item_selector):
                # 单个店铺下的产品
                one_page_result = await parse_shop_page(loaded_page, card_item_selector)
                one_shop_result = one_shop_result + one_page_result

            logger.debug(f'"{one_shop_result.shop_id}" 爬取出 {len(one_shop_result.products)} 个产品')

            # 下载单个店铺的产品图
            await download_one_shop_images(one_shop_result)

            # 保存单个店铺的爬取结果为工作表
            save_sheet(result_workbook, one_shop_result)

            await new_page.close()

    except KeyboardInterrupt:
        logger.warning('KeyboardInterrupt: 爬取任务被中断')
    except asyncio.CancelledError:
        logger.warning('CancelledError: 爬取任务被中断')
    except Exception as e:  # 需要加这个吗？
        logger.error(e)

    finally:
        if len(result_workbook.sheetnames) > 0:
            result_path = write_workbook(
                file=cwd.joinpath('scrape_emag_shop.xlsx'),
                workbook=result_workbook,
                async_mode=False,
            )
            logger.success(f'结果已保存至 "{result_path}"')


def load_shop_workbook() -> Generator[str]:
    """加载店铺表格，每次返回一行店铺链接"""
    # 打开文件选择对话框，选取表格文件
    file = select_file_dialog(title='选择文件', filetypes=[('xlsx file', '.xlsx')])
    logger.info(f'读取工作簿 "{file}"')
    workbook = read_workbook(file=file, async_mode=False, read_only=True)
    sheet_names_str = '"' + '", "'.join(workbook.sheetnames) + '"'

    # 填写表格信息
    while True:
        # 工作表名
        sheet_name = input(f'工作表名（{sheet_names_str}）：')
        # 目标列
        target_column = input('店铺链接所在列（字母）：')
        # 数据开始行号
        start_row = input('从哪行开始（数字）：')
        # 数据结尾行号
        end_row = input('到哪行结束（数字）：')

        # 检查输入结果
        if (
            sheet_name in workbook.sheetnames
            and is_letter(s=target_column)
            and is_number(s=start_row)
            and is_number(s=end_row)
            and int(start_row) <= int(end_row)
        ):
            target_rows = range(int(start_row), int(end_row) + 1)
            break
        else:
            print('输入错误，重新输入')

    sheet = workbook[sheet_name]
    for row in target_rows:
        url = sheet.cell(row=row, column=1).value

        logger.debug(f'读取 {row} 行的 "{url}"')

        if url is None:  # 跳过空行
            continue
        url = str(url)
        if not url.startswith('http'):  # 跳过非链接行
            continue

        yield url


async def wait_page_load(
    page: Page,
    card_item_selector: str,
    target_count: int = 60,
    next_page_timeout: float = 10,
    card_item_timeout: float = 10,
) -> AsyncGenerator[Page]:
    """
    1. 模拟鼠标向下滚动页面，等待加载出足够数量的 card-item 或者时间超时
    2. 传回加载好的页面
    3. 等待加载出下一页按钮
    4. 查看下一页按钮是否可点击，可点击就点击下一页
    """
    while True:
        logger.info(f'等待页面加载 "{page.url}"')

        # 等待加载出足够数量的 card-item 标签
        start_time = perf_counter()
        while True:
            # 数量达标就退出
            card_item_tags = page.locator(card_item_selector)
            card_item_tag_count = await card_item_tags.count()
            if card_item_tag_count >= target_count:
                break

            # 模拟鼠标向下滚动
            await page.mouse.wheel(delta_x=0, delta_y=randint(500, 1000))
            await page.wait_for_timeout(uniform(0, 0.5) * MS1000)

            # 时间超时就退出
            if perf_counter() - start_time >= card_item_timeout:
                break

        # 传回加载好的页面
        yield page

        # 等待下一页的按钮加载出来
        next_page_tag = page.locator('//ul[@id="listing-paginator"]/li[last()]')
        await next_page_tag.wait_for(timeout=next_page_timeout * MS1000, state='attached')

        # 下一页不能点击就退出
        next_page_tag_class_attr = await next_page_tag.get_attribute('class', timeout=MS1000)
        if next_page_tag_class_attr is not None and 'disabled' in next_page_tag_class_attr:
            break

        # 可以点击就点击下一页
        async with page.expect_navigation(timeout=60 * MS1000):
            # 随机等待一会
            await page.wait_for_timeout((10 + randint(0, 10)) * MS1000)
            await next_page_tag.click(timeout=MS1000)


async def parse_shop_page(page: Page, card_item_selector: str) -> Shop:
    """解析商店页面"""
    logger.info(f'解析页面 "{page.url}"')

    # 存储解析的结果
    result = Shop(url=page.url)

    card_item_tags = page.locator(card_item_selector)
    for card_item_tag in await card_item_tags.all():

        # pnk
        data_url: str = await card_item_tag.get_attribute('data-url', timeout=MS1000)  # type: ignore
        pnk_match = re.search(r'/pd/([0-9A-Z]{9})(/|\?|$)', data_url)
        if pnk_match is None:
            continue
        pnk = pnk_match.group(1)

        # Top Favorite
        top_favorite_tag = card_item_tag.locator('//span[text()="Top Favorite"]')
        top_favorite = await top_favorite_tag.count() > 0

        # 产品图链接
        image_url: Optional[str] = None
        image_tag = card_item_tag.locator(
            '//div[@class="img-component position-relative card-v2-thumb-inner"]/img[@src!=""]'
        )
        if await image_tag.count() > 0:
            image_url = await image_tag.get_attribute('src', timeout=MS1000)
            image_url = clean_product_image_url(url=image_url)  # type: ignore

        # 评论数
        review_count: Optional[int] = None
        review_count_tag = card_item_tag.locator(
            '//div[@class="star-rating-text "]/span[@class="visible-xs-inline-block " and text()!=""]'
        )
        if await review_count_tag.count() > 0:
            review_count_text = await review_count_tag.inner_text(timeout=MS1000)
            review_count_text_match = re.search(r'\((\d+)\)', review_count_text)
            if review_count_text_match is not None:
                review_count = int(review_count_text_match.group(1))

        # 价格
        price: Optional[float] = None
        price_tag = card_item_tag.locator('//p[@class="product-new-price"]')
        if await price_tag.count() > 0:
            price_text = await price_tag.inner_text(timeout=MS1000)
            if len(price_text) > 0:
                price = float(re.sub(r'[^\d,]', '', price_text).replace(',', '.'))

        # 单个 card-item 的解析结果
        one_card_item = CardItem(
            pnk=pnk,
            top_favorite=top_favorite,
            thumb_image_url=image_url,
            review_count=review_count,
            price=price,
        )
        result.append(one_card_item)

    logger.debug(f'"{page.url}" 解析到 {len(result.products)} 个产品')
    return result


async def download_one_shop_images(
    shop: Shop,
    concurrent_limit: int = 8,
) -> None:
    """下载单个店铺的产品图"""
    logger.info(f'"{shop.shop_id}" 产品图下载开始')

    # 虚拟的请求头
    fake_header = FakeHttpHeader('chrome').as_header_dict()
    # 并发数限制
    concurrent = asyncio.Semaphore(concurrent_limit)

    # 下载产品图的任务
    async with ClientSession() as client:
        download_imag_tasks = [
            download_image(item.origin_image_url, item.image_save_path, client, fake_header, concurrent)
            for item in shop.products
            if item.image_save_path is not None and item.origin_image_url is not None
        ]
        await asyncio.gather(*download_imag_tasks, return_exceptions=True)

    logger.success(f'"{shop.shop_id}" 产品图下载结束')


async def download_image(
    url: str,
    save_path: Path,
    client: ClientSession,
    headers: dict[str, str],
    semaphore: asyncio.Semaphore,
) -> None:
    """下载单张产品图"""
    # 如果图片已经存在就跳过
    if save_path.exists():
        return

    async with semaphore:
        try:
            logger.info(f'下载产品图 "{url}"')
            async with client.get(url, headers=headers, timeout=ClientTimeout(total=30)) as response:
                response.raise_for_status()
                content = await response.read()
                saved_path = await write_file(file=save_path, data=content, async_mode=True)
                logger.success(f'"{url}" 下载成功，保存至 "{saved_path}"')

        # 发生错误时跳过这张图
        except ClientResponseError as cre:
            logger.error(f'下载产品图时出错\n{cre}')
        except asyncio.CancelledError as ce:
            logger.error(f'下载产品图时出错\n{ce}')
        except TimeoutError as te:
            logger.error(f'下载产品图时出错\n{te}')


def save_sheet(workbook: Workbook, shop: Shop) -> None:
    """保存单个店铺的爬取结果为工作表"""
    sheet: Worksheet = workbook.create_sheet(shop.shop_id)

    # 第一行放店铺链接
    sheet['A1'] = '店铺链接'
    sheet['B1'] = shop.url
    sheet.cell(1, 2).alignment = TEXT_CENTER_WRAP_ALIGNMENT
    sheet.cell(1, 2).hyperlink = Hyperlink(ref=shop.url, target=shop.url)
    sheet.cell(1, 2).font = HYPERLINK_FONT

    # 标题行
    sheet['A2'] = '产品链接'
    sheet['B2'] = '产品图'
    sheet['C2'] = '是否有 Top Favorite'
    sheet['D2'] = '评论数'
    sheet['E2'] = '价格'

    # 设置标题行的样式和各列的宽度
    for col in ['A', 'B', 'C', 'D', 'E']:
        sheet.cell(2, s2i(col)).alignment = TEXT_CENTER_WRAP_ALIGNMENT
        sheet.cell(2, s2i(col)).font = RED_BOLD_FONT
        sheet.cell(2, s2i(col)).fill = YELLOW_FILL
        sheet.column_dimensions[col].width = int(120 / 7)  # 16.04 字符 ≈ 100 磅

    # 遍历爬取结果，插入到工作表中
    for row, item in enumerate(shop.products, start=3):
        # 产品详情页链接
        sheet.cell(row, 1, value=item.url)
        sheet.cell(row, 1).alignment = TEXT_CENTER_WRAP_ALIGNMENT
        sheet.cell(row, 1).hyperlink = Hyperlink(ref=item.url, target=item.url)
        sheet.cell(row, 1).font = HYPERLINK_FONT

        # 产品图
        if item.origin_image_url is not None:
            sheet.cell(row, 2, value=item.origin_image_url)
            sheet.cell(row, 2).alignment = TEXT_CENTER_WRAP_ALIGNMENT
            sheet.cell(row, 2).hyperlink = Hyperlink(ref=item.origin_image_url, target=item.origin_image_url)
            sheet.cell(row, 2).font = HYPERLINK_FONT

            # 检查图片文件是否存在，存在则插入到单元格
            if item.image_save_path is not None and item.image_save_path.exists() and item.image_ext is not None:
                sheet.row_dimensions[row].height = int(120 * 0.75)

                image = read_image(file=item.image_save_path, async_mode=False)
                image = resize_image(image=image, height=120, width=120)
                insert_image(sheet=sheet, image=image, row=row, column=2, image_format=item.image_ext)

        # Top Favorite
        sheet.cell(row, 3, value=str(item.top_favorite))
        sheet.cell(row, 3).alignment = TEXT_CENTER_WRAP_ALIGNMENT

        # 评论数
        if item.review_count is not None:
            sheet.cell(row, 4, value=item.review_count)
            sheet.cell(row, 4).alignment = TEXT_CENTER_WRAP_ALIGNMENT

        # 价格
        if item.price is not None:
            sheet.cell(row, 5, value=item.price)
            sheet.cell(row, 5).alignment = TEXT_CENTER_WRAP_ALIGNMENT

    logger.success(f'"{shop.shop_id}" 的爬取结果保存成功')


if __name__ == '__main__':

    async def main():
        start_time = perf_counter()
        logger.info('程序启动')

        # 创建 chrome_data 文件夹（如果不存在）
        chrome_data_dir = cwd.joinpath('chrome_data')
        if not chrome_data_dir.exists():
            chrome_data_dir.mkdir()

        # 启动浏览器
        abort_res = (ResourceType.IMAGE, ResourceType.MEDIA, ResourceType.FONT)
        browser_manager = PersistentContextManager(
            user_data_dir=chrome_data_dir,
            executable_path=r'C:\Program Files\Google\Chrome\Application\chrome.exe',
            channel='chrome',
            headless=True,
            need_stealth=True,
            abort_res_types=abort_res,
        )
        await browser_manager.start()

        await start_scrape(browser_manager.context)

        # 关闭浏览器
        await browser_manager.close()

        logger.success(f'程序结束，总用时 {round(perf_counter()-start_time, 2)} 秒')

    asyncio.run(main())
