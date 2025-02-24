"""
表格相关工具
"""

from __future__ import annotations

from io import BytesIO as _BytesIO
from typing import TYPE_CHECKING, overload

from openpyxl.reader.excel import load_workbook
from openpyxl.drawing.image import Image as _OpenpyxlImage

from .file_util import read_file as _read_file, write_file as _write_file
from .text_util import is_letter as _is_letter

if TYPE_CHECKING:
    from pathlib import Path
    from typing import Awaitable, Literal

    from openpyxl.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet
    from PIL.Image import Image as PillowImage

    type StrOrPath = str | Path


__all__ = [
    'load_workbook',
    'read_workbook',
    'write_workbook',
    #
    'column_str2int',
    'column_int2str',
    'insert_image',
]

########## 读取 ##########


@overload
async def read_workbook(
    file: StrOrPath,
    async_mode: Literal[True],
    *,
    read_only=False,
    data_only=False,
    keep_links=True,
    rich_text=False,
) -> Workbook:
    """异步读取工作簿"""


@overload
def read_workbook(
    file: StrOrPath,
    async_mode: Literal[False],
    *,
    read_only: bool = False,
    data_only: bool = False,
    keep_links: bool = True,
    rich_text: bool = False,
) -> Workbook:
    """同步读取工作簿"""


def read_workbook(
    file: StrOrPath,
    async_mode: bool,
    *,
    read_only: bool = False,
    data_only: bool = False,
    keep_links: bool = True,
    rich_text: bool = False,
) -> Workbook | Awaitable[Workbook]:
    """读取工作簿"""
    if async_mode:
        return read_workbook_async(
            file=file,
            read_only=read_only,
            data_only=data_only,
            keep_links=keep_links,
            rich_text=rich_text,
        )
    else:
        return read_workbook_sync(
            file=file,
            read_only=read_only,
            data_only=data_only,
            keep_links=keep_links,
            rich_text=rich_text,
        )


async def read_workbook_async(
    file: StrOrPath,
    *,
    read_only: bool = False,
    data_only: bool = False,
    keep_links: bool = True,
    rich_text: bool = False,
) -> Workbook:
    """异步读取工作簿"""
    workbook_bytes = _BytesIO(await _read_file(file=file, mode='bytes', async_mode=True))
    return load_workbook(
        filename=workbook_bytes,
        read_only=read_only,
        data_only=data_only,
        keep_links=keep_links,
        rich_text=rich_text,
    )


def read_workbook_sync(
    file: StrOrPath,
    *,
    read_only: bool = False,
    data_only: bool = False,
    keep_links: bool = True,
    rich_text: bool = False,
) -> Workbook:
    """同步读取工作簿"""
    workbook_bytes = _BytesIO(_read_file(file=file, mode='bytes', async_mode=False))
    return load_workbook(
        filename=workbook_bytes,
        read_only=read_only,
        data_only=data_only,
        keep_links=keep_links,
        rich_text=rich_text,
    )


########## 写入 ##########


@overload
async def write_workbook(file: StrOrPath, workbook: Workbook, async_mode: Literal[True]) -> Path:
    """异步写入工作簿"""


@overload
def write_workbook(file: StrOrPath, workbook: Workbook, async_mode: Literal[False]) -> Path:
    """同步写入工作簿"""


def write_workbook(file: StrOrPath, workbook: Workbook, async_mode: bool) -> Path | Awaitable[Path]:
    """写入工作簿"""
    if async_mode:
        return write_workbook_async(file=file, workbook=workbook)
    else:
        return write_workbook_sync(file=file, workbook=workbook)


async def write_workbook_async(
    file: StrOrPath,
    workbook: Workbook,
) -> Path:
    """异步写入工作簿"""
    workbook_bytes = _BytesIO()
    workbook.save(workbook_bytes)
    return await _write_file(file=file, data=workbook_bytes.getvalue(), async_mode=True)


def write_workbook_sync(
    file: StrOrPath,
    workbook: Workbook,
) -> Path:
    """同步写入工作簿"""
    workbook_bytes = _BytesIO()
    workbook.save(workbook_bytes)
    return _write_file(file=file, data=workbook_bytes.getvalue(), async_mode=False)


########## 列操作 ##########


def column_str2int(column_name: str) -> int:
    """字母形式的列名转成数字形式的列号 A -> 1"""
    if not _is_letter(column_name) or len(column_name) > 3:
        raise ValueError(f'"{column_name}" 不符合列名规范')

    result = 0
    for c in column_name:
        result = result * 26 + (ord(c.upper()) - ord('A') + 1)

    if result > 16384:
        raise ValueError(f'"{column_name}" 超出列名范围 "A" <= column_name <= "XFD"')

    return result


def column_int2str(column_index: int) -> str:
    """数字形式的列号转成字母形式的列名 1 -> A"""
    if 1 <= column_index <= 16384:
        result = ''
        while column_index > 0:
            column_index -= 1
            result = chr(column_index % 26 + ord('A')) + result
            column_index //= 26
        return result
    raise ValueError(f'"{column_index}" 超出列号范围 1 <= column_index <= 16384')


########## 单元格 ##########


def insert_image(
    sheet: Worksheet,
    image: PillowImage,
    row: int,
    column: str | int,
    image_format: str = 'jpeg',
) -> None:
    """往特定单元格插入 Pillow 图片"""
    image_bytes_io = _BytesIO()
    image.save(image_bytes_io, format=image_format)
    column = column if isinstance(column, str) else column_int2str(column_index=column)
    image_bytes_io.seek(0)
    sheet.add_image(_OpenpyxlImage(image_bytes_io), f'{column}{row}')
